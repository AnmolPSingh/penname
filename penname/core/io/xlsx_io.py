"""XLSX path: pseudonymize text cells across all sheets via one shared session.

Formulas pass through verbatim (they are flagged for the user's review in the
GUI, not rewritten), and non-text values are never touched.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from penname.core.detect.crm_templates import header_column_types
from penname.core.engine import PennameSession, RoundTripError
from penname.core.io.limits import check_zip_bomb
from penname.core.replace.applier import reverse_text
from penname.core.types import Mapping, MappingEntry


def _sheet_header(worksheet) -> tuple[dict[int, str], int | None]:
    """Match the first non-empty row against CRM templates. Returns the
    (column-index -> entity type) map and that header row's index, or an empty
    map when the sheet has no confident CRM header."""
    for row in worksheet.iter_rows():
        cells = [c.value if isinstance(c.value, str) else "" for c in row]
        if not any(cells):
            continue
        matches = header_column_types(cells)
        header_index = row[0].row if matches else None
        return matches, header_index
    return {}, None


def _numeric_to_text(value) -> str:
    """Render a numeric cell the way a person would read it (no trailing .0)."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _coerce_like(original, pen_text: str):
    """Keep a numeric cell numeric when the pen value is still a plain number,
    so a pseudonymized amount column stays a number rather than becoming text."""
    if not isinstance(original, (int, float)):
        return pen_text
    try:
        return int(pen_text) if isinstance(original, int) else float(pen_text)
    except ValueError:
        return pen_text


def pseudonymize_xlsx(
    source: str | Path, dest: str | Path, session: PennameSession
) -> Mapping:
    check_zip_bomb(source)  # xlsx is a zip; refuse decompression bombs
    workbook = load_workbook(source)
    entries: dict[tuple[str, str], MappingEntry] = {}
    # Every text cell, changed or not: untouched cells must also survive the
    # union mapping's reversal, or an undetected value could be rewritten into
    # a different donor's data on restore.
    processed: list[tuple[str, str]] = []

    for worksheet in workbook.worksheets:
        column_types, header_row = _sheet_header(worksheet)
        for row in worksheet.iter_rows():
            is_header = header_row is not None and row[0].row == header_row
            for column_index, cell in enumerate(row):
                value = cell.value
                if value is None or is_header:
                    continue
                if isinstance(value, str) and value.startswith("="):
                    continue  # formulas pass through verbatim
                entity_type = column_types.get(column_index)

                if isinstance(value, str):
                    result = (
                        session.pseudonymize_as(value, entity_type)
                        if entity_type is not None
                        else session.pseudonymize(value)
                    )
                    for entry in result.mapping.entries:
                        entries[(entry.entity_type, entry.original)] = entry
                    processed.append((value, result.text))
                    if result.text != value:
                        cell.value = result.text
                elif entity_type is not None and isinstance(value, (int, float)):
                    # A numeric cell in a CRM-named sensitive column (e.g. a
                    # gift amount or constituent ID stored as a number) must be
                    # pseudonymized too, or it silently ships in cleartext.
                    text = _numeric_to_text(value)
                    result = session.pseudonymize_as(text, entity_type)
                    for entry in result.mapping.entries:
                        entries[(entry.entity_type, entry.original)] = entry
                    processed.append((text, result.text))
                    if result.text != text:
                        cell.value = _coerce_like(value, result.text)

    mapping = Mapping(entries=tuple(entries.values()))
    for original, pseudonymized in processed:
        if reverse_text(pseudonymized, mapping) != original:
            raise RoundTripError(
                "a value in this workbook could not be pseudonymized reversibly"
            )

    workbook.save(dest)
    return mapping
