"""CRM column templates: recognize known donor-CRM headers, pseudonymize by column."""

import csv
from pathlib import Path

from penname.core.detect.crm_templates import (
    header_column_types,
    load_column_types,
    match_header,
)
from penname.core.engine import PennameSession
from penname.core.io.csv_io import pseudonymize_csv
from penname.core.replace.applier import reverse_text


def test_templates_load_for_all_five_crms() -> None:
    columns = load_column_types()
    assert columns  # non-empty
    # a representative header from each CRM resolves to the right entity
    assert columns["constituent id"] == "CONSTITUENT_ID"
    assert columns["gift amount"] == "DONATION_AMOUNT"
    assert columns["wealth rating"] == "WEALTH_RATING"


def test_header_matching_is_case_and_punctuation_insensitive() -> None:
    matches = match_header(["Donor_ID", "Gift Amount ", "Notes"])
    assert matches[0] == "CONSTITUENT_ID"
    assert matches[1] == "DONATION_AMOUNT"
    assert 2 not in matches  # "Notes" is not a known column


def test_headerless_file_matches_nothing() -> None:
    # A row of donor data (not column names) must not be treated as a header.
    assert match_header(["Margaret Wilson", "25000", "Cleveland"]) == {}


def test_single_coincidental_match_is_not_treated_as_header() -> None:
    """A headerless data row where one cell happens to equal a CRM header word
    must NOT be treated as a header (that would skip the row and leak it)."""
    row = ["Margaret Wilson", "Annual Fund", "25000"]  # only "Fund"-ish coincidence
    assert header_column_types(row) == {}  # below the majority threshold


def test_confident_header_passes_threshold() -> None:
    row = ["Constituent ID", "Gift Amount", "Notes"]  # 2 of 3 match
    matched = header_column_types(row)
    assert matched[0] == "CONSTITUENT_ID" and matched[1] == "DONATION_AMOUNT"


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def test_crm_column_forces_id_and_amount_even_when_opaque(tmp_path: Path) -> None:
    """An opaque ID column and a bare-number amount column get pseudonymized
    because the header names them, not because the values look sensitive."""
    source = tmp_path / "export.csv"
    _write_csv(
        source,
        [
            ["Donor ID", "Gift Amount", "Fund"],
            ["88213", "25000", "CLEANWATER"],
            ["88214", "1500", "GENERAL"],
        ],
    )
    out = tmp_path / "export.pen.csv"

    mapping = pseudonymize_csv(source, out, PennameSession())

    with open(out, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    assert rows[0] == ["Donor ID", "Gift Amount", "Fund"]  # header untouched
    # opaque id and bare-number amount both replaced
    assert rows[1][0] != "88213"
    assert rows[1][1] != "25000"
    types = {e.entity_type for e in mapping.entries}
    assert "CONSTITUENT_ID" in types and "DONATION_AMOUNT" in types

    # everything still reverses
    for orig_row, pen_row in zip(
        list(csv.reader(open(source, newline="", encoding="utf-8"))), rows
    ):
        for orig_cell, pen_cell in zip(orig_row, pen_row):
            assert reverse_text(pen_cell, mapping) == orig_cell


def test_crm_column_applies_to_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    from penname.core.io.xlsx_io import pseudonymize_xlsx

    wb = Workbook()
    ws = wb.active
    ws.append(["Constituent ID", "Gift Amount", "Fund"])
    ws.append(["88213", 25000, "CLEANWATER"])
    ws.append(["88214", 1500, "GENERAL"])
    source = tmp_path / "export.xlsx"
    wb.save(source)
    out = tmp_path / "export.pen.xlsx"

    mapping = pseudonymize_xlsx(source, out, PennameSession())

    pen = load_workbook(out)["Sheet"]
    assert [c.value for c in pen[1]] == ["Constituent ID", "Gift Amount", "Fund"]
    assert pen["A2"].value != "88213"  # opaque id forced by column
    # numeric amount in a CRM column is pseudonymized too, and stays numeric
    assert pen["B2"].value != 25000
    assert isinstance(pen["B2"].value, (int, float))
    types = {e.entity_type for e in mapping.entries}
    assert "CONSTITUENT_ID" in types and "DONATION_AMOUNT" in types


def test_crm_column_consistency(tmp_path: Path) -> None:
    """The same ID in two rows gets the same pen name via the column path."""
    source = tmp_path / "export.csv"
    _write_csv(
        source,
        [["Donor ID"], ["88213"], ["88213"]],
    )
    out = tmp_path / "out.csv"
    pseudonymize_csv(source, out, PennameSession())

    rows = list(csv.reader(open(out, newline="", encoding="utf-8")))
    assert rows[1][0] == rows[2][0]
    assert rows[1][0] != "88213"
