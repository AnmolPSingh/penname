"""XLSX path: pseudonymize text cells across all sheets; formulas and numbers untouched."""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from penname.core.engine import PennameSession
from penname.core.io.xlsx_io import pseudonymize_xlsx
from penname.core.replace.applier import reverse_text


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Donors"
    ws.append(["Full Name", "Email", "Last Gift", "Notes"])
    ws.append(["Margaret Wilson", "m.wilson@homemail.com", 25000, "Met at the gala in Cleveland."])
    ws.append(["Robert Castellano", "r.castellano88@gmail.com", 1500, "Prefers phone calls."])
    ws.append(["Margaret Wilson", "m.wilson@homemail.com", 300, ""])
    ws["C5"] = "=SUM(C2:C4)"

    notes = wb.create_sheet("Meeting Notes")
    notes["A1"] = "Angela Ruiz met Robert Castellano in Santa Fe on June 3, 2025."

    path = tmp_path / "donors.xlsx"
    wb.save(path)
    return path


def test_xlsx_cells_round_trip(workbook_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "donors.pen.xlsx"
    mapping = pseudonymize_xlsx(workbook_path, out, PennameSession())

    original = load_workbook(workbook_path)
    pen = load_workbook(out)
    assert original.sheetnames == pen.sheetnames

    for sheet_name in original.sheetnames:
        for orig_row, pen_row in zip(
            original[sheet_name].iter_rows(), pen[sheet_name].iter_rows()
        ):
            for orig_cell, pen_cell in zip(orig_row, pen_row):
                if isinstance(orig_cell.value, str):
                    assert reverse_text(str(pen_cell.value), mapping) == orig_cell.value
                else:
                    assert pen_cell.value == orig_cell.value  # numbers untouched


def test_xlsx_replaces_names_and_preserves_formulas(
    workbook_path: Path, tmp_path: Path
) -> None:
    out = tmp_path / "donors.pen.xlsx"
    pseudonymize_xlsx(workbook_path, out, PennameSession())

    pen = load_workbook(out)
    donors = pen["Donors"]
    values = [str(c.value) for row in donors.iter_rows() for c in row]
    assert "Margaret Wilson" not in values
    assert donors["C5"].value == "=SUM(C2:C4)"  # formula preserved verbatim
    # repeated donor still consistent across rows
    assert donors["A2"].value == donors["A4"].value


def test_xlsx_consistency_spans_sheets(workbook_path: Path, tmp_path: Path) -> None:
    out = tmp_path / "donors.pen.xlsx"
    mapping = pseudonymize_xlsx(workbook_path, out, PennameSession())

    pens = {e.original: e.pen_name for e in mapping.entries}
    assert "Robert Castellano" in pens
    pen = load_workbook(out)
    assert pens["Robert Castellano"] in str(pen["Meeting Notes"]["A1"].value)
    assert pens["Robert Castellano"] == str(pen["Donors"]["A3"].value)
